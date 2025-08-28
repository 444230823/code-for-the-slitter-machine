import numpy as np
import matplotlib.pyplot as plt
from vmdpy import VMD
from PyEMD import EMD,CEEMDAN,EEMD
import openpyxl
from sklearn.metrics import mean_absolute_error,mean_squared_error,root_mean_squared_error
from scipy.stats import pearsonr
from pykalman import KalmanFilter

wb=openpyxl.load_workbook(r'VMD分解-1次审稿补实验2.xlsx',data_only=True)
ws=wb['0716-1-6圈'] #工作表
f_list=[]
VGXR_list=[]
t=[]
R_com=[]#合成信号
R_com_value=0#合成值
R_loss=0
R_loss_last=0
rsa=[[],[],[],[]]#记录r(s,a)
# -----测试信号及其参数--start-------------
for each,i in zip(ws['B2':'B123'],range(1,124)):
    for cell in each:
        f_list.append(cell.value)
        t.append(i)
# for each,i in zip(ws1['E8':'E123'],range(1,118)):
#     for cell in each:
#         VGXR_list.append(cell.value)
f=np.array(f_list)

uniform_noise=np.array(np.random.uniform(-0.005,0.005,122))#定义噪声:范围最小值，范围最大值，个数
gaussian_noise=np.random.normal(0,0.01,122)      #定义噪声:均值，标准差，size
poisson_noise=np.random.poisson(1560,122)/100000 #定义噪声:lambda=1560
f+=uniform_noise#+噪声


# uniform_noise=np.array(np.random.uniform(-0.01,0.01,116))#定义噪声:范围最小值，范围最大值，个数
# VGXR_list+=uniform_noise
Fs=1000 # 采样频率
N=62  # 采样点数

# t=np.arange(1,N+1)/N
t=np.array(t)


fre_axis=np.linspace(0,Fs/2,int(N/2))
f_1=100;f_2=200;f_3=300
v_1=(np.cos(2*np.pi*f_1*t));v_2=1/4*(np.cos(2*np.pi*f_2*t));v_3=1/16*(np.cos(2*np.pi*f_3*t))
v=[v_1,v_2,v_3] # 测试信号所包含的各成分
# f=v_1+v_2+v_3+0.1*np.random.randn(v_1.size)  # 测试信号
# -----测试信号及其参数--end----------
# alpha 惩罚系数；带宽限制经验取值为抽样点长度1.5-2.0倍.
# 惩罚系数越小，各IMF分量的带宽越大，过大的带宽会使得某些分量包含其他分量言号;
# a值越大，各IMF分量的带宽越小，过小的带宽是使得被分解的信号中某些信号丢失该系数常见取值范围为1000~3000
alpha=1400
# alpha=0
tau=0   # tau 噪声容限，即允许重构后的信号与原始信号有差别。
K=5    # K 分解模态（IMF）个数
DC=0    # DC 若为0则让第一个IMF为直流分量/趋势向量
init=1  # init 指每个IMF的中心频率进行初始化。当初始化为1时，进行均匀初始化。
tol=1e-7# 控制误差大小常量，决定精度与迭代次数
u,u_hat,omega=VMD(f,alpha,tau,K,DC,init,tol) # 输出U是各个IMF分量，u_hat是各IMF的频谱，omega为各IMF的中心频率
emd=EMD.EMD();imfs_emd=emd(f)#imfs_emd[-1]是滤波值
eemd=EEMD.EEMD(ext_EMD=emd);imfs_eemd=eemd(f)#imfs_eemd[-1]是滤波值
ceemd=CEEMDAN.CEEMDAN(ext_EMD=emd);imfs_ceemd=ceemd(f)#imfs_ceemd[-1]是滤波值
kf=KalmanFilter(initial_state_mean=0.29,n_dim_obs=1)#Kalman filter
imfs_Kalman,_=kf.filter(f)#Kalman filter
imfs_Kalman=imfs_Kalman.flatten()#把shape是(122, 1)的array转换为shape是(122, )的array
high_frequency=[];low_frequency=[]

for i in range(len(u.T)):
    R_com_value=0
    for j in range(len(u.T[0])):
        R_com_value+=u.T[i][j]
    R_com.append(R_com_value)
    high_frequency.append(u.T[i][0])
    low_frequency.append(u.T[i][1])


# plt.figure(figsize=(10,7));plt.subplot(K+1,1,1);plt.plot(t,f)
# plt.figure(figsize=(10,7));plt.plot(t,u.T);plt.title('all Decomposed modes');plt.show()  # u.T是对u的转置
for each,i in zip(ws['C2':'C123'],range(len(low_frequency))):
    for cell in each:
        cell.value=high_frequency[i]
# wb.save('VMD分解-1次审稿补实验2.xlsx')

rmse=mean_squared_error(f,high_frequency)#均方差
cc=np.corrcoef(f,high_frequency)#相关系数
snr=10*np.log10(np.mean(np.array(high_frequency)**2)/np.mean(np.array(low_frequency)**2))
print('VMD indexes','\n',rmse,'\n',cc,'\n',snr)

rmse=mean_squared_error(f,imfs_emd[-1])#均方差
cc=np.corrcoef(f,imfs_emd[-1])#相关系数
snr=10*np.log10(np.mean(np.array(imfs_emd[-1])**2)/np.mean(np.array(imfs_emd[-3])**2))
print('EMD indexes','\n',rmse,'\n',cc,'\n',snr)

rmse=mean_squared_error(f,imfs_eemd[-1])#均方差
cc=np.corrcoef(f,imfs_eemd[-1])#相关系数
snr=10*np.log10(np.mean(np.array(imfs_eemd[-1])**2)/np.mean(np.array(imfs_eemd[-2])**2))
print('EEMD indexes','\n',rmse,'\n',cc,'\n',snr)

rmse=mean_squared_error(f,imfs_ceemd[-1])#均方差
cc=np.corrcoef(f,imfs_ceemd[-1])#相关系数
snr=10*np.log10(np.mean(np.array(imfs_ceemd[-1])**2)/np.mean(np.array(imfs_ceemd[-4])**2))
print('CEEMD indexes','\n',rmse,'\n',cc,'\n',snr)

rmse=mean_squared_error(f,imfs_Kalman)#均方差
cc=np.corrcoef(f,imfs_Kalman)#相关系数
snr=10*np.log10(np.mean(np.array(imfs_Kalman)**2)/np.mean(np.array(f-imfs_Kalman)**2))
print('Kalman indexes','\n',rmse,'\n',cc,'\n',snr)